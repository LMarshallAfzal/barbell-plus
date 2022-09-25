import csv
import os
import openpyxl
import xlrd

path = "/home/leonard/projects/barbell-plus/barbellplus/assets/programs/"
filename1 = "KCL Barbell Program GPP Leonard 20210614.xlsm"
filename2 = "002 KCL Barbell Program Phase 2A  Barbell Introduction.xlsm"
filename3 = "KCL Barbell Program Leonard  Basic Strength V1.xlsm"
filename4 = "KCL Barbell Program Leonard 3 Week Volume Block 20210705.xlsm"
filename5 = "KCL Barbell Program Post Lockdown Leonard 20210412.xlsm"

book = openpyxl.load_workbook(f"{path}{filename5}")

sheet1 = book["Training Plan"]
sheet2 = book["Exercise Selection"]

session_1 = []
session_2 = []
session_3 = []
session_4 = []

num_exercises_in_session_1 = 0
for i in range(2, 10):
    if sheet2.cell(row=i, column=1).value is None:
        break
    num_exercises_in_session_1 += 1
num_exercises_in_session_2 = 0
for i in range(2, 10):
    if sheet2.cell(row=i, column=2).value is None:
        break
    num_exercises_in_session_2 += 1
num_exercises_in_session_3 = 0
for i in range(2, 10):
    if sheet2.cell(row=i, column=3).value is None:
        break
    num_exercises_in_session_3 += 1
num_exercises_in_session_4 = 0
for i in range(2, 10):
    if sheet2.cell(row=i+14, column=1).value is None:
        break
    num_exercises_in_session_4 += 1

for i in range(0, num_exercises_in_session_1):
    session_1.append({"exercise": '', "sets": [], "reps": [], "rest": []})
for i in range(0, num_exercises_in_session_2):
    session_2.append({"exercise": '', "sets": [], "reps": [], "rest": []})
for i in range(0, num_exercises_in_session_3):
    session_3.append({"exercise": '', "sets": [], "reps": [], "rest": []})
for i in range(0, num_exercises_in_session_4):
    session_4.append({"exercise": '', "sets": [], "reps": [], "rest": []})

weeks = 0
for i in range(0, 12):
    if sheet1.cell(row=i+7, column=2).value.startswith("Week"):
        weeks += 1
    else:
        break

increaser = 8
cell_ref = ''
starting_row = 4 
for i in range(0, num_exercises_in_session_1):
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row, min_col=2, max_col=2):
        for cell in row:
            cell_ref = cell.value
            cell_ref = cell_ref.removeprefix("='Exercise Selection'!")
            exercise = sheet2[cell_ref].value
            session_1[i]["exercise"] = exercise[4:]
            starting_row += 3
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=3, max_col=3):
        for cell in row:
            session_1[i]["sets"].append(cell.value)
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=4, max_col=4):
        for cell in row:
            session_1[i]["reps"].append(cell.value)
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=6, max_col=6):
        for cell in row:
            session_1[i]["rest"].append(cell.value)
    starting_row += increaser

starting_row = 100
for i in range(0, num_exercises_in_session_2):
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row, min_col=2, max_col=2):
        for cell in row:
            cell_ref = cell.value
            cell_ref = cell_ref.removeprefix("='Exercise Selection'!")
            exercise = sheet2[cell_ref].value
            session_2[i]["exercise"] = exercise[4:]
            starting_row += 3
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=3, max_col=3):
        for cell in row:
            session_2[i]["sets"].append(cell.value)
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=4, max_col=4):
        for cell in row:
            session_2[i]["reps"].append(cell.value)
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=6, max_col=6):
        for cell in row:
            session_2[i]["rest"].append(cell.value)
    starting_row += increaser

starting_row = 196
for i in range(0, num_exercises_in_session_3):
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row, min_col=2, max_col=2):
        for cell in row:
            cell_ref = cell.value
            cell_ref = cell_ref.removeprefix("='Exercise Selection'!")
            exercise = sheet2[cell_ref].value
            session_3[i]["exercise"] = exercise[4:]
            starting_row += 3
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=3, max_col=3):
        for cell in row:
            session_3[i]["sets"].append(cell.value)
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=4, max_col=4):
        for cell in row:
            session_3[i]["reps"].append(cell.value)
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=6, max_col=6):
        for cell in row:
            session_3[i]["rest"].append(cell.value)
    starting_row += increaser

starting_row = 100
for i in range(0, num_exercises_in_session_4):
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row, min_col=2, max_col=2):
        for cell in row:
            cell_ref = cell.value
            cell_ref = cell_ref.removeprefix("='Exercise Selection'!")
            exercise = sheet2[cell_ref].value
            session_4[i]["exercise"] = exercise[4:]
            starting_row += 3
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=3, max_col=3):
        for cell in row:
            session_4[i]["sets"].append(cell.value)
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=4, max_col=4):
        for cell in row:
            session_4[i]["reps"].append(cell.value)
    for row in sheet1.iter_rows(min_row=starting_row, max_row=starting_row+weeks-1, min_col=6, max_col=6):
        for cell in row:
            session_4[i]["rest"].append(cell.value)
    starting_row += increaser

print(session_1)
print(session_2)
print(session_3)
print(session_4)

file = filename5.replace('.xlsm', '.csv')
file = file.replace(' ', '-').lower()
with open(file, 'w', newline='') as csvfile:
    fieldnames = ['exercise', 'sets', 'reps', 'rest']
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()
    for i in range(0, num_exercises_in_session_1):
        writer.writerow(session_1[i])
    writer.writerow({})
    for i in range(0, num_exercises_in_session_2):
        writer.writerow(session_2[i])
    writer.writerow({})
    for i in range(0, num_exercises_in_session_3):
        writer.writerow(session_3[i])
    writer.writerow({})
    for i in range(0, num_exercises_in_session_4):
        writer.writerow(session_4[i])
    writer.writerow({})